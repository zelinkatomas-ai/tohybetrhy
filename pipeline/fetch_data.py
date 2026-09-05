#!/usr/bin/env python3
"""
Datová pipeline: stáhne ceny ETF (Yahoo Finance chart API), spočítá momentum
v NATIVNÍ měně fondu a vygeneruje JSON soubory pro web.

Momentum měříme v měně fondu – kurzové riziko je téma samo o sobě, ale pro
určení toho, "co jede", by přepočet měn jen přidával šum.

Výstupy (src/data/):
  - regions.json         ... regiony (tabulka + graf hlavních čtyř)
  - sectors.json         ... americké sektory, SPDR (tabulka, řazeno dle 3M)
  - crypto.json          ... Bitcoin (tabulka + graf)
  - momentum_etfs.json   ... momentum ETF vs benchmarky (tabulka + graf)
  - smart_money.json     ... chytré peníze vs retail (CoT, NAAIM, pákové ETF)
  - social.json          ... sítě: Reddit (ApeWisdom + tradestie), StockTwits
  - liquidity.json       ... likvidita: čistá likvidita Fedu, peníze vs inflace,
                             zaparkovaná hotovost, trh vs M2, dolar
  - risks.json           ... rizika: index GPR, nejistota EPU, obranný sektor
  - polymarket.json      ... sázkové trhy: geopolitika + makro (top dle objemu)
  - summary.json         ... generovaný slovní vzkaz pro hlavní stránku

Spouštění:  python3 pipeline/fetch_data.py
Závislosti: requests (pip install -r pipeline/requirements.txt)
"""

from __future__ import annotations

import json
import time
from datetime import date, datetime, timezone
from pathlib import Path

import requests

# ---------------------------------------------------------------------------
# Konfigurace – JEDINÉ místo, kde se přidávají/mění sledované tituly.
#   chart=True      ... titul se kreslí v čárovém grafu skupiny
#   benchmark=True  ... v grafu čárkovaně, v tabulce oddělen jako srovnání
# ---------------------------------------------------------------------------
CHART_START = "2024-01-01"   # všechny grafy začínají tady
HISTORY_RANGE = "4y"          # stahovaná historie (kvůli 12M oknu a 52t průměru)
MOMENTUM_WEEKS = {"r1": 4, "r3": 13, "r6": 26, "r12": 52}
SMA_WEEKS = 52

GROUPS: dict[str, dict] = {
    "regions": {
        "file": "regions.json",
        "note": "Regiony přes ETF, výnosy v měně fondu. Hlavní čtyři regiony i v grafu.",
        "items": [
            {"ticker": "IWDA.AS", "name": "Svět (MSCI World)",   "chart": True, "core": True},
            {"ticker": "CSPX.AS", "name": "USA (S&P 500)",       "chart": True, "core": True},
            {"ticker": "MEUD.PA", "name": "Evropa (STOXX 600)",  "chart": True, "core": True},
            {"ticker": "EMIM.AS", "name": "Rozvíjející se trhy", "chart": True, "core": True},
            {"ticker": "MCHI",    "name": "Čína"},
            {"ticker": "QDV5.DE", "name": "Indie"},
            {"ticker": "CSJP.MI", "name": "Japonsko"},
            {"ticker": "ILF",     "name": "Jižní Amerika"},
            {"ticker": "EZA",     "name": "Afrika (JAR)"},
        ],
    },
    "sectors": {
        "file": "sectors.json",
        "note": "Americké sektory (Select Sector SPDR), výnosy v USD. Řazeno podle 3M momenta.",
        "sort_by": "r3",
        "items": [
            {"ticker": "XLK",  "name": "Technologie"},
            {"ticker": "XLC",  "name": "Komunikační služby"},
            {"ticker": "XLY",  "name": "Zbytná spotřeba"},
            {"ticker": "XLF",  "name": "Finance"},
            {"ticker": "XLV",  "name": "Zdravotnictví"},
            {"ticker": "XLI",  "name": "Průmysl"},
            {"ticker": "XLB",  "name": "Materiály"},
            {"ticker": "XLE",  "name": "Energie"},
            {"ticker": "XLP",  "name": "Základní spotřeba"},
            {"ticker": "XLU",  "name": "Utility"},
            {"ticker": "XLRE", "name": "Reality"},
        ],
    },
    "regions_europe": {
        "file": "regions_europe.json",
        "note": "Evropa v detailu po zemích. U menších trhů americká iShares ETF jako indikátor. Výnosy v měně fondu.",
        "items": [
            {"ticker": "EXS1.DE", "name": "Německo (DAX)"},
            {"ticker": "CAC.PA",  "name": "Francie (CAC 40)"},
            {"ticker": "ISF.L",   "name": "Velká Británie (FTSE 100)"},
            {"ticker": "EWL",     "name": "Švýcarsko"},
            {"ticker": "EWI",     "name": "Itálie"},
            {"ticker": "EPOL",    "name": "Polsko"},
            # Česko: likvidní ETF neexistuje a index PX není na Yahoo
            # (ověřeno 2026-09-05) – viz TODO, Známé věci
        ],
        "sort_by": "r3",
    },
    "regions_asia": {
        "file": "regions_asia.json",
        "note": "Asie v detailu po zemích; Čína, Indie a Japonsko mají řádky v hlavní tabulce regionů. Americká iShares/VanEck ETF jako indikátor, výnosy v USD.",
        "items": [
            {"ticker": "EWT",  "name": "Tchaj-wan"},
            {"ticker": "EWY",  "name": "Jižní Korea"},
            {"ticker": "EWS",  "name": "Singapur"},
            {"ticker": "VNM",  "name": "Vietnam"},
            {"ticker": "EIDO", "name": "Indonésie"},
        ],
        "sort_by": "r3",
    },
    "crypto": {
        "file": "crypto.json",
        "note": "Bitcoin a Ethereum, spotové ceny v USD. Obchodují se nonstop; vzorkujeme týdně jako ostatní data.",
        "items": [
            {"ticker": "BTC-USD", "name": "Bitcoin", "chart": True},
            {"ticker": "ETH-USD", "name": "Ethereum", "chart": True},
        ],
    },
    "momentum_etfs": {
        "file": "momentum_etfs.json",
        "note": "ETF stavěná na momentum faktoru vs široký trh (čárkovaně). Výnosy v měně fondu.",
        "items": [
            {"ticker": "SPMO",    "name": "SPMO – S&P 500 Momentum",            "chart": True},
            {"ticker": "MTUM",    "name": "MTUM – MSCI USA Momentum",           "chart": True},
            {"ticker": "XSMO",    "name": "XSMO – S&P SmallCap Momentum",       "chart": True},
            {"ticker": "XMMO",    "name": "XMMO – S&P MidCap Momentum",         "chart": True},
            {"ticker": "FMTM",    "name": "FMTM – Focused U.S. Momentum",       "chart": True},
            {"ticker": "IDMO",    "name": "IDMO – Intl Developed Momentum",     "chart": True},
            {"ticker": "IEMO.MI", "name": "IEMO – Europe Momentum (UCITS)"},
            {"ticker": "PIE",     "name": "PIE – Emerging Markets Momentum"},
            {"ticker": "EEMO",    "name": "EEMO – S&P EM Momentum"},
            {"ticker": "SPY",     "name": "USA (S&P 500)",   "chart": True, "benchmark": True},
            {"ticker": "URTH",    "name": "Svět (MSCI World)", "chart": True, "benchmark": True},
        ],
    },
}

OUT_DIR = Path(__file__).resolve().parent.parent / "src" / "data"
HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; momentum-web-pipeline)"}


def fetch_yahoo_weekly(ticker: str) -> tuple[dict[str, float], str]:
    """Vrátí ({ISO datum: adjusted close}, měna)."""
    url = (
        f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}"
        f"?range={HISTORY_RANGE}&interval=1wk&events=div%2Csplit"
    )
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    result = r.json()["chart"]["result"][0]
    currency = result["meta"].get("currency", "?")
    ts = result["timestamp"]
    ind = result["indicators"]
    closes = (ind.get("adjclose", [{}])[0].get("adjclose")
              or ind["quote"][0]["close"])
    out: dict[str, float] = {}
    for t, c in zip(ts, closes):
        if c is not None:
            out[date.fromtimestamp(t).isoformat()] = round(float(c), 4)
    return out, currency


def pct(series: list[float], weeks_back: int) -> float | None:
    if len(series) <= weeks_back:
        return None
    old, new = series[-1 - weeks_back], series[-1]
    return round((new / old - 1) * 100, 2)


def build_group(key: str, cfg: dict) -> None:
    rows = []
    raw_series = []  # {name, benchmark, points: {date: indexovaná hodnota}}

    for item in cfg["items"]:
        print(f"[{key}] Stahuji {item['ticker']} ({item['name']}) ...")
        try:
            prices, currency = fetch_yahoo_weekly(item["ticker"])
        except Exception as e:  # jeden mrtvý ticker nesmí shodit celou skupinu
            print(f"[{key}] {item['ticker']} SELHALO: {e}")
            continue
        time.sleep(1)  # ohleduplnost k API

        dates = sorted(prices)
        vals = [prices[d] for d in dates]

        sma = (sum(vals[-SMA_WEEKS:]) / SMA_WEEKS) if len(vals) >= SMA_WEEKS else None
        rows.append({
            "ticker": item["ticker"],
            "name": item["name"],
            "currency": currency,
            "benchmark": bool(item.get("benchmark")),
            "core": bool(item.get("core")),
            **{k: pct(vals, w) for k, w in MOMENTUM_WEEKS.items()},
            "signal": ("above" if vals[-1] >= sma else "below") if sma else None,
        })

        if item.get("chart"):
            idx = [i for i, d in enumerate(dates) if d >= CHART_START]
            base = vals[idx[0]]
            raw_series.append({
                "name": item["name"],
                "benchmark": bool(item.get("benchmark")),
                "points": {dates[i]: round(vals[i] / base * 100, 2) for i in idx},
            })

    # řazení tabulky (benchmarky vždy dole)
    if cfg.get("sort_by"):
        k = cfg["sort_by"]
        rows.sort(key=lambda r: (r["benchmark"], -(r[k] if r[k] is not None else -999)))

    # graf: společná osa, série zarovnané přes null
    chart = None
    if raw_series:
        chart_dates = sorted({d for s in raw_series for d in s["points"]})
        chart = {
            "dates": chart_dates,
            "series": [
                {"name": s["name"], "benchmark": s["benchmark"],
                 "values": [s["points"].get(d) for d in chart_dates]}
                for s in raw_series
            ],
        }

    (OUT_DIR / cfg["file"]).write_text(json.dumps({
        "updated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "note": cfg["note"],
        "rows": rows,
        "chart": chart,
    }, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"[{key}] -> {cfg['file']}")


# ---------------------------------------------------------------------------
# Chytré peníze vs retail
# ---------------------------------------------------------------------------

def fetch_cot() -> dict:
    """CFTC Traders in Financial Futures – E-mini S&P 500, čisté pozice
    jako % open interestu. Asset manažeři = instituce, leveraged funds =
    hedge fondy, nonreportable = malí obchodníci (proxy retailu)."""
    import urllib.parse
    where = urllib.parse.quote(
        f"contract_market_name = 'E-MINI S&P 500' AND "
        f"report_date_as_yyyy_mm_dd >= '{CHART_START}'"
    )
    url = ("https://publicreporting.cftc.gov/resource/gpe5-46if.json"
           f"?$where={where}&$order=report_date_as_yyyy_mm_dd&$limit=5000")
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    dates, inst, hedge, small = [], [], [], []
    for row in r.json():
        oi = float(row["open_interest_all"])
        if oi <= 0:
            continue
        net = lambda l, s: round((float(row[l]) - float(row[s])) / oi * 100, 2)
        dates.append(row["report_date_as_yyyy_mm_dd"][:10])
        inst.append(net("asset_mgr_positions_long", "asset_mgr_positions_short"))
        hedge.append(net("lev_money_positions_long", "lev_money_positions_short"))
        small.append(net("nonrept_positions_long_all", "nonrept_positions_short_all"))
    return {
        "dates": dates,
        "series": [
            {"name": "Instituce (asset manažeři)", "values": inst},
            {"name": "Hedge fondy (leveraged funds)", "values": hedge},
            {"name": "Malí obchodníci (retail)", "values": small},
        ],
    }


def fetch_naaim() -> dict:
    """NAAIM Exposure Index – průměrná akciová expozice aktivních správců
    (0 = mimo trh, 100 = plně zainvestováno).

    Primárně hledáme xlsx s historií na stránce programu (NAAIM ho občas
    publikuje, občas ne); záložně čteme data z jejich vloženého grafu
    (index.naaim.org/embeddable/chart). Pokud je poslední hodnota starší
    než 60 dní, indikátor raději vynecháme, než abychom ukazovali
    zastaralá čísla."""
    import html as html_mod
    import io
    import re

    points: dict[str, float] = {}

    # 1) xlsx s kompletní historií (pokud je odkaz na stránce)
    try:
        page = requests.get("https://naaim.org/programs/naaim-exposure-index/",
                            headers=HEADERS, timeout=30)
        m = re.search(r'href="([^"]+\.xlsx)"', page.text)
        if m:
            from openpyxl import load_workbook
            xlsx = requests.get(m.group(1), headers=HEADERS, timeout=60)
            wb = load_workbook(io.BytesIO(xlsx.content), read_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            header = next(rows)
            i_date = header.index("Date")
            i_mean = next(i for i, h in enumerate(header) if h and "Mean" in str(h))
            for row in rows:
                d, v = row[i_date], row[i_mean]
                if d is None or v is None:
                    continue
                iso = d.date().isoformat() if hasattr(d, "date") else str(d)[:10]
                if iso >= CHART_START:
                    points[iso] = round(float(v), 1)
    except Exception as e:
        print(f"[smart_money] NAAIM xlsx nedostupné ({e}), zkouším embed")

    # 2) záloha: data z vloženého grafu
    if not points:
        t = requests.get("https://index.naaim.org/embeddable/chart",
                         headers=HEADERS, timeout=30).text
        u = html_mod.unescape(t)
        lab = re.search(r'"labels":(\[[^\]]*\])', u)
        dat = re.search(r'"data":(\[[^\]]*\])', u)
        if not (lab and dat):
            raise RuntimeError("NAAIM: data nenalezena ani v embedu")
        labels = json.loads(lab.group(1))
        values = json.loads(dat.group(1))
        for d, v in zip(labels, values):
            if str(d)[:10] >= CHART_START:
                points[str(d)[:10]] = round(float(v), 1)

    dates = sorted(points)
    if not dates:
        raise RuntimeError("NAAIM: žádná data")
    # kontrola čerstvosti – zastaralý indikátor je horší než žádný
    age_days = (date.today() - date.fromisoformat(dates[-1])).days
    if age_days > 60:
        raise RuntimeError(f"NAAIM: poslední hodnota je {age_days} dní stará, vynechávám")
    return {
        "dates": dates,
        "series": [{"name": "NAAIM Exposure Index", "values": [points[d] for d in dates]}],
    }


def fetch_retail_proxy() -> dict:
    """Proxy retailového apetitu: podíl pákových ETF (TQQQ, SQQQ, SPXL, SPXS)
    na dolarovém objemu vůči SPY + QQQ. Pákové ETF obchoduje převážně retail."""
    def dollar_volume(ticker: str) -> dict[str, float]:
        url = (f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}"
               f"?range={HISTORY_RANGE}&interval=1wk")
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
        res = r.json()["chart"]["result"][0]
        q = res["indicators"]["quote"][0]
        out = {}
        for t, c, v in zip(res["timestamp"], q["close"], q["volume"]):
            if c and v:
                out[date.fromtimestamp(t).isoformat()] = c * v
        return out

    lev = ["TQQQ", "SQQQ", "SPXL", "SPXS"]
    base = ["SPY", "QQQ"]
    vols: dict[str, dict[str, float]] = {}
    for t in lev + base:
        print(f"[smart_money] Stahuji objemy {t} ...")
        vols[t] = dollar_volume(t)
        time.sleep(1)

    common = sorted(set.intersection(*(set(v) for v in vols.values())))
    common = [d for d in common if d >= CHART_START]
    values = [
        round(sum(vols[t][d] for t in lev) / sum(vols[t][d] for t in base) * 100, 1)
        for d in common
    ]
    return {
        "dates": common,
        "series": [{"name": "Objem pákových ETF vs SPY+QQQ (%)", "values": values}],
    }


def fetch_reddit() -> dict:
    """Reddit sentiment ze dvou veřejných API bez klíče:

    - ApeWisdom (primární): top tickery podle počtu zmínek za 24 h napříč
      investičními subreddity (WSB, r/stocks…) + upvoty.
    - Tradestie (doplněk + záloha): směr sentimentu z analýzy komentářů na
      WSB (bullish/bearish) – ApeWisdom měří jen hlasitost, ne směr. Když
      ApeWisdom vypadne, sestaví se žebříček z tradestie (počet komentářů).

    Změna t/t se počítá proti snapshotu z minulého běhu pipeline uloženému
    přímo v smart_money.json – jen pokud snapshot pochází ze stejného zdroje
    (zmínky ApeWisdom a komentáře tradestie nejsou srovnatelné)."""
    import html as html_mod

    # tradestie: {ticker: {"sentiment": "bullish"/"bearish", "score": float, ...}}
    tradestie: dict[str, dict] = {}
    try:
        r = requests.get("https://tradestie.com/api/v1/apps/reddit",
                         headers=HEADERS, timeout=30)
        r.raise_for_status()
        for row in r.json():
            tradestie[row["ticker"]] = {
                "sentiment": (row.get("sentiment") or "").lower() or None,
                "score": row.get("sentiment_score"),
                "comments": int(row.get("no_of_comments") or 0),
            }
    except Exception as e:
        print(f"[smart_money] tradestie nedostupné ({e}), pojedeme bez sentimentu")

    # ApeWisdom -> jednotný tvar [(rank, ticker, name, mentions, upvotes)]
    source, rows = "apewisdom", []
    try:
        r = requests.get("https://apewisdom.io/api/v1.0/filter/all-stocks/page/1",
                         headers=HEADERS, timeout=30)
        r.raise_for_status()
        rows = [(int(x["rank"]), x["ticker"], html_mod.unescape(x["name"]),
                 int(x["mentions"]), int(x["upvotes"]))
                for x in r.json()["results"]]
    except Exception as e:
        print(f"[smart_money] ApeWisdom nedostupné ({e}), záloha: tradestie")
        source = "tradestie"
        ranked = sorted(tradestie.items(), key=lambda kv: -kv[1]["comments"])
        rows = [(i + 1, t, t, v["comments"], None)
                for i, (t, v) in enumerate(ranked)]
    if not rows:
        raise RuntimeError("Reddit: ApeWisdom ani tradestie nevrátily data")

    # předchozí snapshot (kvůli změně t/t) – čteme starý soubor před přepisem;
    # fallback na smart_money.json kvůli migraci (reddit tam bydlel dřív)
    prev_mentions: dict[str, int] = {}
    prev_date: str | None = None
    for prev_file in ("social.json", "smart_money.json"):
        try:
            old = json.loads((OUT_DIR / prev_file).read_text(encoding="utf-8"))
            snap = (old.get("reddit") or {}).get("snapshot") or {}
            if snap and snap.get("source", "apewisdom") == source:  # jen srovnatelné zdroje
                prev_mentions = snap.get("mentions") or {}
                prev_date = snap.get("date")
            if snap:
                break
        except Exception:
            pass

    top = []
    for rank, ticker, name, mentions, upvotes in rows[:10]:
        prev = prev_mentions.get(ticker)
        sent = tradestie.get(ticker) or {}
        top.append({
            "rank": rank,
            "ticker": ticker,
            "name": name,
            "mentions": mentions,
            "upvotes": upvotes,
            "prev_mentions": prev,
            "change_pct": round((mentions / prev - 1) * 100, 1) if prev else None,
            "sentiment": sent.get("sentiment"),
            "sentiment_score": sent.get("score"),
        })

    return {
        "top": top,
        "prev_date": prev_date,
        "source": source,
        # snapshot top 50, aby měly změnu i tituly, které se do top 10 teprve dostanou
        "snapshot": {
            "date": date.today().isoformat(),
            "source": source,
            "mentions": {t: m for _, t, _, m, _ in rows[:50]},
        },
    }


def fetch_vix() -> dict:
    """VIX – implikovaná volatilita S&P 500, tedy cena pojistky proti
    poklesu. K tomu termínová struktura VIX3M/VIX: normálně nad 1
    (tříměsíční ochrana dražší než okamžitá, contango); pod 1 znamená,
    že trh platí za okamžitou ochranu víc než za budoucí – akutní stres
    (backwardace). Obojí týdně z Yahoo, jako ostatní ceny."""
    vix, _ = fetch_yahoo_weekly("^VIX")
    time.sleep(1)
    vix3m, _ = fetch_yahoo_weekly("^VIX3M")

    dates_all = sorted(vix)
    vals = [vix[d] for d in dates_all]
    sma = [
        round(sum(vals[i - SMA_WEEKS + 1:i + 1]) / SMA_WEEKS, 2)
        if i >= SMA_WEEKS - 1 else None
        for i in range(len(vals))
    ]
    keep = [i for i, d in enumerate(dates_all) if d >= CHART_START]
    common = [d for d in dates_all if d in vix3m and d >= CHART_START]
    return {
        "level": {
            "dates": [dates_all[i] for i in keep],
            "series": [
                {"name": "VIX", "values": [round(vals[i], 2) for i in keep]},
                {"name": "52týdenní průměr", "values": [sma[i] for i in keep], "benchmark": True},
            ],
        },
        "term": {
            "dates": common,
            "series": [{"name": "VIX3M / VIX",
                        "values": [round(vix3m[d] / vix[d], 3) for d in common]}],
        },
    }


def fetch_stocktwits() -> dict:
    """StockTwits trending: symboly, o kterých se na platformě právě nejvíc
    mluví (jejich trending algoritmus), plus počet sledujících (watchers)
    jako měřítko trvalejší popularity. Veřejné API bez klíče."""
    r = requests.get("https://api.stocktwits.com/api/2/trending/symbols.json",
                     headers=HEADERS, timeout=30)
    r.raise_for_status()
    return {
        "top": [
            {"rank": i + 1, "ticker": s["symbol"], "name": s["title"],
             "watchers": int(s.get("watchlist_count") or 0)}
            for i, s in enumerate(r.json()["symbols"][:10])
        ],
    }


def build_smart_money() -> None:
    charts = {}
    for key, fn in [("cot", fetch_cot), ("naaim", fetch_naaim),
                    ("retail", fetch_retail_proxy), ("vix", fetch_vix)]:
        try:
            print(f"[smart_money] {key} ...")
            charts[key] = fn()
        except Exception as e:  # jeden rozbitý zdroj nesmí shodit celý build
            print(f"[smart_money] {key} SELHALO: {e}")
            charts[key] = None

    (OUT_DIR / "smart_money.json").write_text(json.dumps({
        "updated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "note": "CoT: čisté pozice v E-mini S&P 500 futures jako % open interestu (CFTC, týdně). "
                "NAAIM: průměrná akciová expozice aktivních správců. "
                "Retail proxy: podíl pákových ETF na dolarovém objemu. "
                "VIX: implikovaná volatilita S&P 500 + termínová struktura VIX3M/VIX.",
        **charts,
    }, ensure_ascii=False, indent=1), encoding="utf-8")
    print("[smart_money] -> smart_money.json")


def build_social() -> None:
    """Sekce Sítě: o čem se mluví na investičních sociálních sítích."""
    charts = {}
    for key, fn in [("reddit", fetch_reddit), ("stocktwits", fetch_stocktwits)]:
        try:
            print(f"[social] {key} ...")
            charts[key] = fn()
        except Exception as e:  # jeden rozbitý zdroj nesmí shodit celý build
            print(f"[social] {key} SELHALO: {e}")
            charts[key] = None

    (OUT_DIR / "social.json").write_text(json.dumps({
        "updated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "note": "Reddit: zmínky za 24 h dle ApeWisdom, změna proti předchozí aktualizaci, "
                "směr sentimentu (bullish/bearish) dle tradestie. "
                "StockTwits: trending symboly + počet sledujících. Sítě se aktualizují denně.",
        **charts,
    }, ensure_ascii=False, indent=1), encoding="utf-8")
    print("[social] -> social.json")


# ---------------------------------------------------------------------------
# Likvidita – hromadí se v systému volná hotovost?
# Zdroje: FRED (CSV bez klíče), ECB Data Portal (CSV bez klíče), Yahoo Finance.
# ---------------------------------------------------------------------------

YOY_START = "2023-01-01"  # měsíční série tahame o rok dřív kvůli meziročním změnám


# FRED z datacenter IP (GitHub Actions) často blokuje ne-browser klienty
# (request visí do timeoutu), proto plné browser hlavičky + retry. Záloha:
# oficiální API, pokud je v prostředí bezplatný klíč FRED_API_KEY
# (https://fred.stlouisfed.org/docs/api/api_key.html, secret v GitHub Actions).
FRED_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Accept": "text/csv,text/plain,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://fred.stlouisfed.org/",
}


def _fred_api(series_id: str, start: str, api_key: str) -> dict[str, float]:
    r = requests.get(
        "https://api.stlouisfed.org/fred/series/observations",
        params={"series_id": series_id, "api_key": api_key,
                "file_type": "json", "observation_start": start},
        headers=HEADERS, timeout=30)
    r.raise_for_status()
    return {obs["date"]: float(obs["value"])
            for obs in r.json()["observations"] if obs["value"] not in (".", "")}


def fetch_fred_csv(series_id: str, start: str = CHART_START) -> dict[str, float]:
    """Série z FRED. S klíčem (FRED_API_KEY) jde první oficiální API –
    veřejný CSV endpoint z IP GitHub runnerů visí do timeoutu, takže by
    jen zdržoval; bez klíče zkoušíme CSV (funguje z běžných IP).
    Vrací {ISO datum: hodnota}; chybějící pozorování (".") vynechává."""
    import csv
    import io
    import os

    out: dict[str, float] = {}
    api_key = os.environ.get("FRED_API_KEY")
    if api_key:
        try:
            out = _fred_api(series_id, start, api_key)
        except Exception as e:
            print(f"[liquidity] FRED {series_id} API selhalo: {e}")

    if not out:
        url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}&cosd={start}"
        for attempt in (1, 2):
            try:
                r = requests.get(url, headers=FRED_HEADERS, timeout=30)
                r.raise_for_status()
                rows = csv.reader(io.StringIO(r.text))
                next(rows)  # hlavička: datum + id série (jméno sloupce se měnilo)
                for row in rows:
                    if len(row) >= 2 and row[1] not in (".", ""):
                        out[row[0][:10]] = float(row[1])
                break
            except Exception as e:
                print(f"[liquidity] FRED {series_id} CSV pokus {attempt} selhal: {e}")
                time.sleep(2)

    if not out:
        raise RuntimeError(f"FRED {series_id}: nedostupné"
                           + ("" if api_key else " (API klíč není nastaven)"))
    return out


def to_trillions(points: dict[str, float]) -> dict[str, float]:
    """Převod na biliony USD. FRED udává různé série v milionech, jiné
    v miliardách; jednotku poznáme z řádu hodnot (TGA/RRP/bilance Fedu
    se reálně pohybují v řádu stovek miliard až jednotek bilionů USD)."""
    peak = max(abs(v) for v in points.values())
    scale = 1e6 if peak >= 1e5 else 1e3 if peak >= 1e2 else 1
    return {d: v / scale for d, v in points.items()}


def fetch_ecb_csv(series_key: str, start: str) -> dict[str, float]:
    """Série z ECB Data Portal (SDMX-CSV, bez API klíče). Měsíční periody
    „2024-01" normalizuje na první den měsíce, ať sedí k FRED datům."""
    import csv
    import io

    flow, key = series_key.split(".", 1)
    url = (f"https://data-api.ecb.europa.eu/service/data/{flow}/{key}"
           f"?format=csvdata&startPeriod={start}")
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    out: dict[str, float] = {}
    for row in csv.DictReader(io.StringIO(r.text)):
        period, val = row.get("TIME_PERIOD"), row.get("OBS_VALUE")
        if period and val:
            out[period if len(period) == 10 else f"{period}-01"] = float(val)
    if not out:
        raise RuntimeError(f"ECB {series_key}: prázdná odpověď")
    return out


def yoy(points: dict[str, float]) -> dict[str, float]:
    """Meziroční změna v % z měsíční série klíčované {YYYY-MM-01: hodnota}."""
    return {
        d: round((v / points[prev] - 1) * 100, 2)
        for d, v in points.items()
        if (prev := f"{int(d[:4]) - 1}{d[4:]}") in points and points[prev]
    }


def monthly_last(points: dict[str, float]) -> dict[str, float]:
    """Denní/týdenní sérii převede na měsíční (poslední pozorování v měsíci)."""
    out: dict[str, float] = {}
    for d in sorted(points):
        out[d[:7] + "-01"] = points[d]
    return out


def fetch_net_liquidity() -> dict:
    """Čistá likvidita Fedu = bilance − účet ministerstva financí (TGA)
    − reverzní repo (ON RRP). Peníze v TGA a RRP jsou z trhu fakticky
    odčerpané; zbytek jsou volné rezervy bank – palivo pro riziková aktiva."""
    fed = to_trillions(fetch_fred_csv("WALCL"))      # bilance Fedu, týdně (středa)
    tga = to_trillions(fetch_fred_csv("WTREGEN"))    # pokladna, týdně (středa)
    rrp = to_trillions(fetch_fred_csv("RRPONTSYD"))  # reverzní repo, denně

    def last_upto(points: dict[str, float], d: str) -> float | None:
        cands = [k for k in points if k <= d]
        return points[max(cands)] if cands else None

    dates, net, bal = [], [], []
    for d in sorted(fed):
        t, rr = last_upto(tga, d), last_upto(rrp, d)
        if t is None or rr is None:
            continue
        dates.append(d)
        bal.append(round(fed[d], 2))                              # bil. USD
        net.append(round(fed[d] - t - rr, 2))
    return {
        "dates": dates,
        "series": [
            {"name": "Čistá likvidita", "values": net},
            {"name": "Bilance Fedu celkem", "values": bal, "benchmark": True},
        ],
    }


def fetch_money_vs_inflation_us() -> dict:
    """USA: růst peněžní zásoby M2 vs inflace CPI, obojí meziročně. Kladná
    mezera = reálné peněžní zásoby přibývá (přebytečná likvidita)."""
    m2 = yoy(fetch_fred_csv("M2SL", YOY_START))
    cpi = yoy(fetch_fred_csv("CPIAUCSL", YOY_START))
    dates = sorted(d for d in m2 if d in cpi and d >= CHART_START)
    return {
        "dates": dates,
        "series": [
            {"name": "Peněžní zásoba M2 (meziročně)", "values": [m2[d] for d in dates]},
            {"name": "Inflace CPI (meziročně)", "values": [cpi[d] for d in dates], "benchmark": True},
        ],
    }


def fetch_eurostat_hicp(start: str) -> dict[str, float]:
    """Meziroční inflace HICP pro eurozónu přímo od Eurostatu (JSON-stat,
    bez klíče). Eurostat je primární zdroj HICP – řady u ECB zamrzly na
    2025-12 po rebasi indexu na 2025=100. Zkoušíme starý i kandidátní nový
    kód datasetu po rebasi a řady sloučíme (meziroční míra na bázi nezávisí)."""
    out: dict[str, float] = {}
    for dataset in ("prc_hicp_manr", "prc_hicp25_manr"):
        url = ("https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/"
               f"data/{dataset}?format=JSON&lang=EN&coicop=CP00&geo=EA"
               f"&sinceTimePeriod={start}")
        try:
            r = requests.get(url, headers=HEADERS, timeout=30)
            r.raise_for_status()
            js = r.json()
            # jediná proměnná dimenze je čas -> pozice v "value" = index času
            idx = js["dimension"]["time"]["category"]["index"]
            vals = js["value"]
            got = {f"{p}-01": float(vals[str(i)]) for p, i in idx.items() if str(i) in vals}
            print(f"[liquidity] Eurostat {dataset}: do {max(got) if got else '–'}")
            out.update(got)
        except Exception as e:
            print(f"[liquidity] Eurostat {dataset} nedostupný: {e}")
    if not out:
        raise RuntimeError("Eurostat HICP: žádná data")
    return out


def fetch_money_vs_inflation_ea() -> dict:
    """Eurozóna: totéž s M3 (ECB Data Portal) a HICP (Eurostat). Osa jede
    podle M3; pokud HICP po rebasi na 2025=100 zaostává, jeho čára končí
    dřív (poctivější než osekávat obojí na průnik)."""
    m3_raw = fetch_ecb_csv("BSI.M.U2.Y.V.M30.X.1.U2.2300.Z01.E", YOY_START[:7])
    hicp = fetch_eurostat_hicp(CHART_START[:7])  # už meziroční
    print(f"[liquidity] money_ea: M3 do {max(m3_raw)}, HICP do {max(hicp)}")
    m3 = yoy(m3_raw)
    dates = sorted(d for d in m3 if d >= CHART_START)
    return {
        "dates": dates,
        "series": [
            {"name": "Peněžní zásoba M3 (meziročně)", "values": [m3[d] for d in dates]},
            {"name": "Inflace HICP (meziročně)", "values": [hicp.get(d) for d in dates], "benchmark": True},
        ],
    }


def fetch_cash_parked() -> dict:
    """Kde se hromadí hotovost: bankovní vklady (celý sektor) a retailové
    fondy peněžního trhu (suchý prach investorů), meziroční růst. K tomu
    podíl retail MMF na M2 – kolik peněžní zásoby sedí „na parkovišti"."""
    dep = monthly_last(fetch_fred_csv("DPSACBW027SBOG", YOY_START))  # vklady, týdně
    # RMFSL = měsíční retail MMF; týdenní WRMFSL skončila s přechodem H.6
    # na měsíční frekvenci (2021)
    mmf = fetch_fred_csv("RMFSL", YOY_START)
    m2 = fetch_fred_csv("M2SL", YOY_START)
    dep_yoy, mmf_yoy = yoy(dep), yoy(mmf)

    mmf_t, m2_t = to_trillions(mmf), to_trillions(m2)  # kvůli podílu srovnat jednotky
    dates = sorted(d for d in dep_yoy if d in mmf_yoy and d >= CHART_START)
    share = {d: round(mmf_t[d] / m2_t[d] * 100, 2) for d in mmf_t if d in m2_t}
    sdates = sorted(share)
    return {
        "dates": dates,
        "series": [
            {"name": "Vklady v bankách (meziročně)", "values": [dep_yoy[d] for d in dates]},
            {"name": "Retail fondy peněžního trhu (meziročně)", "values": [mmf_yoy[d] for d in dates]},
        ],
        "mmf_share": {
            "last": share[sdates[-1]],
            "prev_year": share[sdates[-13]] if len(sdates) >= 13 else None,
        },
    }


def fetch_market_vs_m2() -> dict:
    """Trh vs peněžní zásoba: S&P 500 dělený M2, indexováno. Dlouhé okno
    (od 2015) – jde o strukturální ukazatel, ne o momentum. Čárkovaně
    pětiletý klouzavý průměr poměru jako trend."""
    url = ("https://query1.finance.yahoo.com/v8/finance/chart/%5EGSPC"
           "?range=15y&interval=1mo")
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    res = r.json()["chart"]["result"][0]
    closes = res["indicators"]["quote"][0]["close"]
    spx = {}
    for t, c in zip(res["timestamp"], closes):
        if c is not None:
            spx[date.fromtimestamp(t).isoformat()[:7] + "-01"] = float(c)

    m2 = fetch_fred_csv("M2SL", "2010-01-01")
    common = sorted(d for d in spx if d in m2)
    base = spx[common[0]] / m2[common[0]]
    ratio = [round(spx[d] / m2[d] / base * 100, 1) for d in common]
    # trend: klouzavý průměr přes posledních až 60 měsíců (min. 36 kvůli rozjezdu)
    trend = [
        round(sum(ratio[max(0, i - 59):i + 1]) / len(ratio[max(0, i - 59):i + 1]), 1)
        if i >= 35 else None
        for i in range(len(ratio))
    ]
    keep = [i for i, d in enumerate(common) if d >= "2015-01-01"]
    return {
        "dates": [common[i] for i in keep],
        "series": [
            {"name": "S&P 500 / M2 (index)", "values": [ratio[i] for i in keep]},
            {"name": "Pětiletý průměr poměru", "values": [trend[i] for i in keep], "benchmark": True},
        ],
    }


def fetch_dollar() -> dict:
    """Dolarový index jako teploměr globální likvidity: silný dolar = utažené
    dolarové financování ve světě, slabý dolar = uvolněné."""
    prices, _ = fetch_yahoo_weekly("DX-Y.NYB")
    dates_all = sorted(prices)
    vals = [prices[d] for d in dates_all]
    sma = [
        round(sum(vals[i - SMA_WEEKS + 1:i + 1]) / SMA_WEEKS, 2)
        if i >= SMA_WEEKS - 1 else None
        for i in range(len(vals))
    ]
    keep = [i for i, d in enumerate(dates_all) if d >= CHART_START]
    return {
        "dates": [dates_all[i] for i in keep],
        "series": [
            {"name": "Dolarový index (DXY)", "values": [round(vals[i], 2) for i in keep]},
            {"name": "52týdenní průměr", "values": [sma[i] for i in keep], "benchmark": True},
        ],
    }


def build_liquidity() -> None:
    charts = {}
    for key, fn in [("net_liquidity", fetch_net_liquidity),
                    ("money_us", fetch_money_vs_inflation_us),
                    ("money_ea", fetch_money_vs_inflation_ea),
                    ("cash", fetch_cash_parked),
                    ("market_m2", fetch_market_vs_m2),
                    ("dollar", fetch_dollar)]:
        try:
            print(f"[liquidity] {key} ...")
            charts[key] = fn()
            time.sleep(1)
        except Exception as e:  # jeden rozbitý zdroj nesmí shodit celý build
            print(f"[liquidity] {key} SELHALO: {e}")
            charts[key] = None

    (OUT_DIR / "liquidity.json").write_text(json.dumps({
        "updated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "note": "Čistá likvidita: bilance Fedu − TGA − ON RRP (bil. USD, týdně). "
                "Peníze vs inflace: M2/M3 a CPI/HICP meziročně (FRED, ECB). "
                "Hotovost: vklady a retail MMF meziročně, podíl MMF na M2. "
                "Trh vs M2: S&P 500 / M2, index. Dolar: DXY vs 52t průměr.",
        **charts,
    }, ensure_ascii=False, indent=1), encoding="utf-8")
    print("[liquidity] -> liquidity.json")


# ---------------------------------------------------------------------------
# Rizika – kvantifikovaná geopolitika a nejistota
# ---------------------------------------------------------------------------

def _read_table(content: bytes) -> list[list]:
    """Načte tabulku z xlsx/xls/CSV podle skutečného obsahu souboru
    (přípona na akademických webech nemusí odpovídat formátu)."""
    import io

    if content[:2] == b"PK":  # xlsx (zip)
        from openpyxl import load_workbook
        ws = load_workbook(io.BytesIO(content), read_only=True).active
        return [list(r) for r in ws.iter_rows(values_only=True)]
    if content[:4] == b"\xd0\xcf\x11\xe0":  # starý binární .xls
        import xlrd
        from xlrd.xldate import xldate_as_datetime
        book = xlrd.open_workbook(file_contents=content)
        sh = book.sheet_by_index(0)
        return [
            [xldate_as_datetime(sh.cell(r, c).value, book.datemode)
             if sh.cell(r, c).ctype == xlrd.XL_CELL_DATE else sh.cell(r, c).value
             for c in range(sh.ncols)]
            for r in range(sh.nrows)
        ]
    import csv  # poslední záchrana: prostý CSV
    return [r for r in csv.reader(io.StringIO(content.decode("utf-8", "replace")))]


def weekly_mean(points: dict[str, float]) -> dict[str, float]:
    """Denní sérii zprůměruje po týdnech (klíč = pondělí daného týdne) –
    denní hodnoty novinových indexů jsou příliš rozeskákané."""
    from datetime import timedelta

    buckets: dict[str, list[float]] = {}
    for d, v in points.items():
        day = date.fromisoformat(d)
        monday = day - timedelta(days=day.weekday())
        buckets.setdefault(monday.isoformat(), []).append(v)
    return {k: sum(vs) / len(vs) for k, vs in buckets.items()}


def rolling_year(wk: dict[str, float]) -> dict[str, float]:
    """Roční klouzavý průměr týdenní série – „na co si trhy zvykly
    za poslední rok". Eskalace se pozná proti němu, ne proti průměru
    za celé dekády."""
    weeks = sorted(wk)
    return {
        k: sum(wk[w] for w in weeks[i - 51:i + 1]) / 52
        for i, k in enumerate(weeks) if i >= 51
    }


def fetch_gpr() -> dict:
    """Geopolitical Risk Index (Caldara–Iacoviello, Fed): podíl novinových
    článků o geopolitickém napětí v 10 velkých denících. Akademický
    standard; autoři publikují denní řadu zdarma na svém webu."""
    content = None
    for url in ("https://www.matteoiacoviello.com/gpr_files/data_gpr_daily_recent.xls",
                "https://www.matteoiacoviello.com/gpr_files/data_gpr_daily_recent.xlsx"):
        try:
            r = requests.get(url, headers=HEADERS, timeout=60)
            r.raise_for_status()
            content = r.content
            break
        except Exception as e:
            print(f"[risks] GPR {url.rsplit('/', 1)[-1]} nedostupný: {e}")
    if content is None:
        raise RuntimeError("GPR: soubor nedostupný")

    rows = _read_table(content)
    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    i_date = next(i for i, h in enumerate(header) if h in ("date", "day", "yyyymmdd"))
    i_val = next(i for i, h in enumerate(header) if h == "gprd")

    points: dict[str, float] = {}
    for row in rows[1:]:
        d, v = row[i_date], row[i_val]
        if d is None or v in (None, ""):
            continue
        iso = d.date().isoformat() if hasattr(d, "date") else str(d).strip()[:10]
        if iso[:8].isdigit() and len(iso.strip()) == 8:  # tvar YYYYMMDD
            iso = f"{iso[:4]}-{iso[4:6]}-{iso[6:8]}"
        try:
            points[iso] = float(v)
        except (TypeError, ValueError):
            continue
    if not points:
        raise RuntimeError("GPR: v souboru nejsou čitelná data")

    wk = weekly_mean(points)
    roll = rolling_year(wk)
    avg = sum(points.values()) / len(points)
    dates = sorted(d for d in wk if d >= CHART_START)
    return {
        "dates": dates,
        "series": [
            {"name": "GPR (týdenní průměr)", "values": [round(wk[d], 1) for d in dates]},
            {"name": "Roční klouzavý průměr", "values": [round(roll[d], 1) if d in roll else None for d in dates]},
            {"name": f"Průměr od {min(points)[:4]}", "values": [round(avg, 1)] * len(dates),
             "benchmark": True},
        ],
        "avg": round(avg, 1),
        "avg_1y": round(roll[dates[-1]], 1) if dates and dates[-1] in roll else None,
    }


def fetch_epu() -> dict:
    """Economic Policy Uncertainty (Baker–Bloom–Davis): nejistota
    z hospodářské politiky (cla, sankce, rozpočty) z novinových textů.
    Denní řada pro USA je přímo na FREDu."""
    daily = fetch_fred_csv("USEPUINDXD", "2015-01-01")
    wk = weekly_mean(daily)
    roll = rolling_year(wk)
    avg = sum(daily.values()) / len(daily)
    dates = sorted(d for d in wk if d >= CHART_START)
    return {
        "dates": dates,
        "series": [
            {"name": "EPU (týdenní průměr)", "values": [round(wk[d], 1) for d in dates]},
            {"name": "Roční klouzavý průměr", "values": [round(roll[d], 1) if d in roll else None for d in dates]},
            {"name": "Průměr od 2015", "values": [round(avg, 1)] * len(dates), "benchmark": True},
        ],
        "avg": round(avg, 1),
        "avg_1y": round(roll[dates[-1]], 1) if dates and dates[-1] in roll else None,
    }


def fetch_defense() -> dict:
    """Trh hlasuje o zbrojení: relativní síla obranných ETF vůči S&P 500
    (poměr cen, indexováno). Trvalý růst poměru = trh oceňuje strukturální
    geopolitické riziko bez ohledu na titulky."""
    spy, _ = fetch_yahoo_weekly("SPY")
    time.sleep(1)
    ita, _ = fetch_yahoo_weekly("ITA")    # iShares U.S. Aerospace & Defense
    time.sleep(1)
    euad, _ = fetch_yahoo_weekly("EUAD")  # Select STOXX Europe Aerospace & Defense

    def rel(etf: dict[str, float]) -> dict[str, float]:
        common = [d for d in sorted(etf) if d in spy and d >= CHART_START]
        if not common:
            return {}
        base = etf[common[0]] / spy[common[0]]
        return {d: round(etf[d] / spy[d] / base * 100, 1) for d in common}

    us, eu = rel(ita), rel(euad)
    dates = sorted(set(us) | set(eu))
    return {
        "dates": dates,
        "series": [
            {"name": "USA: obrana vs S&P 500 (ITA)", "values": [us.get(d) for d in dates]},
            {"name": "Evropa: obrana vs S&P 500 (EUAD)", "values": [eu.get(d) for d in dates]},
        ],
    }


# tagy Polymarketu -> naše skupiny; výběr otázek je mechanický (top objem),
# žádná ruční kurátorská volba, takže se seznam sám obměňuje s děním
POLYMARKET_GROUPS: dict[str, set[str]] = {
    "geopolitics": {"geopolitics", "world", "war", "ukraine", "russia", "israel",
                    "middle-east", "china", "iran", "nato", "north-korea", "taiwan"},
    "macro": {"economy", "macro", "fed", "fed-rates", "interest-rates", "rates",
              "inflation", "recession", "tariffs", "trade-war", "treasury"},
}


def fetch_polymarket() -> dict:
    """Sázkové trhy Polymarket: pravděpodobnosti konkrétních otázek oceněné
    penězi sázkařů (veřejné gamma API bez klíče). Bereme nejobchodovanější
    otevřené otázky za 24 h a třídíme je podle tagů do skupin – geopolitika
    pro Rizika, makro (sazby, recese, cla) pro Likviditu."""
    r = requests.get("https://gamma-api.polymarket.com/events",
                     params={"closed": "false", "limit": "300",
                             "order": "volume24hr", "ascending": "false"},
                     headers=HEADERS, timeout=30)
    r.raise_for_status()
    events = r.json()
    if isinstance(events, dict):  # některé verze API balí seznam do obálky
        events = events.get("events") or events.get("data") or []

    def num(x) -> float:
        try:
            return float(x)
        except (TypeError, ValueError):
            return 0.0

    def parsed(m: dict) -> dict | None:
        o, p = m.get("outcomes"), m.get("outcomePrices")
        if isinstance(o, str):
            o = json.loads(o)
        if isinstance(p, str):
            p = json.loads(p)
        if not o or not p:
            return None
        probs = [num(x) for x in p]
        # binární trh: bereme cenu "Yes"; jinak nejpravděpodobnější výstup
        try:
            i = [str(x).lower() for x in o].index("yes")
        except ValueError:
            i = probs.index(max(probs))
        return {"question": m.get("question"), "outcome": str(o[i]), "prob": probs[i]}

    rows = []
    for ev in events:
        try:
            # otázky s prošlým deadlinem jen čekají na vypořádání – vynechat
            end = str(ev.get("endDate") or "")
            if end and end[:10] < date.today().isoformat():
                continue
            tags = {str(t.get("slug") or t.get("label") or "").strip().lower()
                    for t in (ev.get("tags") or [])}
            cands = [c for c in (parsed(m) for m in (ev.get("markets") or [])
                                 if not m.get("closed")) if c]
            if not cands:
                continue
            # jediný trh = binární otázka; u vícetrhových událostí (volby
            # apod.) ukazujeme lídra, ne nejobchodovanější dílčí trh
            best = cands[0] if len(cands) == 1 else max(cands, key=lambda c: c["prob"])
            if best["prob"] >= 0.995:  # prakticky rozhodnuto, čeká na vypořádání
                continue
            rows.append({
                "question": (ev.get("title") if len(cands) == 1 else best["question"])
                            or ev.get("title") or best["question"],
                "outcome": best["outcome"],
                "prob": round(best["prob"] * 100, 1),
                "volume24h": round(num(ev.get("volume24hr")), 0),
                "url": f"https://polymarket.com/event/{ev['slug']}" if ev.get("slug") else None,
                "_tags": tags,
            })
        except Exception:
            continue  # jedna rozbitá otázka nesmí shodit celý výběr

    rows.sort(key=lambda x: -x["volume24h"])
    groups: dict[str, list] = {}
    for gname, gtags in POLYMARKET_GROUPS.items():
        groups[gname] = [{k: v for k, v in r_.items() if k != "_tags"}
                         for r_ in rows if r_["_tags"] & gtags][:6]
    if not any(groups.values()):
        raise RuntimeError(f"Polymarket: z {len(rows)} otázek žádná neprošla filtrem tagů")
    print(f"[polymarket] {len(rows)} otázek, "
          + ", ".join(f"{k}: {len(v)}" for k, v in groups.items()))
    return groups


def build_polymarket() -> None:
    try:
        groups = fetch_polymarket()
    except Exception as e:
        print(f"[polymarket] SELHALO: {e}")
        groups = {k: None for k in POLYMARKET_GROUPS}

    (OUT_DIR / "polymarket.json").write_text(json.dumps({
        "updated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "note": "Polymarket: pravděpodobnosti oceněné penězi sázkařů. Výběr mechanicky – "
                "nejobchodovanější otevřené otázky (objem 24 h) s tagy dané skupiny. "
                "Týdenní snímek; živé kurzy na polymarket.com.",
        **groups,
    }, ensure_ascii=False, indent=1), encoding="utf-8")
    print("[polymarket] -> polymarket.json")


def build_risks() -> None:
    charts = {}
    for key, fn in [("gpr", fetch_gpr), ("epu", fetch_epu), ("defense", fetch_defense)]:
        try:
            print(f"[risks] {key} ...")
            charts[key] = fn()
            time.sleep(1)
        except Exception as e:  # jeden rozbitý zdroj nesmí shodit celý build
            print(f"[risks] {key} SELHALO: {e}")
            charts[key] = None

    (OUT_DIR / "risks.json").write_text(json.dumps({
        "updated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "note": "GPR: novinový index geopolitického rizika (Caldara–Iacoviello), týdenní průměr. "
                "EPU: nejistota z hospodářské politiky (FRED). "
                "Obrana: relativní síla obranných ETF vůči S&P 500.",
        **charts,
    }, ensure_ascii=False, indent=1), encoding="utf-8")
    print("[risks] -> risks.json")


# ---------------------------------------------------------------------------
# Generovaný slovní vzkaz pro hlavní stránku
# ---------------------------------------------------------------------------

def _fmt(v: float) -> str:
    return f"{v:+.1f}".replace(".", ",").replace("+", "+") + " %"


def build_summary() -> None:
    regions = json.loads((OUT_DIR / "regions.json").read_text(encoding="utf-8"))
    sectors = json.loads((OUT_DIR / "sectors.json").read_text(encoding="utf-8"))
    etfs = json.loads((OUT_DIR / "momentum_etfs.json").read_text(encoding="utf-8"))
    smart = json.loads((OUT_DIR / "smart_money.json").read_text(encoding="utf-8"))
    crypto = json.loads((OUT_DIR / "crypto.json").read_text(encoding="utf-8"))

    by_ticker = {r["ticker"]: r for r in regions["rows"]}

    # 1) globální vzkaz podle světového indexu
    world = by_ticker["IWDA.AS"]
    r3, above = world["r3"] or 0, world["signal"] == "above"
    if above and r3 >= 2:
        s_global = (f"Globální momentum zůstává pozitivní – světový index se drží nad "
                    f"52týdenním průměrem a za poslední tři měsíce přidal {_fmt(r3)}.")
    elif above and r3 >= 0:
        s_global = ("Globální momentum je neutrální až mírně pozitivní – světový index "
                    "se drží nad 52týdenním průměrem, tempo růstu ale polevilo.")
    elif above:
        s_global = (f"Globální momentum slábne – trh se zatím drží nad dlouhodobým "
                    f"průměrem, poslední tři měsíce jsou ale ztrátové ({_fmt(r3)}).")
    elif r3 > 0:
        s_global = (f"Trh se odráží ode dna – světový index je stále pod 52týdenním "
                    f"průměrem, krátkodobé momentum ale sílí ({_fmt(r3)} za tři měsíce).")
    else:
        s_global = (f"Globální momentum je negativní – světový index je pod 52týdenním "
                    f"průměrem a za tři měsíce ztrácí {_fmt(r3)}.")

    # 2) rotace mezi hlavními regiony podle 4týdenního výnosu
    decl = {  # skloňování pro věty
        "CSPX.AS": ("USA", "USA"),
        "MEUD.PA": ("Evropy", "Evropě"),
        "EMIM.AS": ("rozvíjejících se trhů", "rozvíjejícím se trhům"),
    }
    ranked = sorted(decl, key=lambda t: by_ticker[t]["r1"] or 0, reverse=True)
    spread = (by_ticker[ranked[0]]["r1"] or 0) - (by_ticker[ranked[-1]]["r1"] or 0)
    if spread < 1.5:
        s_rotation = ("Rozdíly mezi hlavními regiony jsou v posledních čtyřech týdnech "
                      "malé – výrazná rotace kapitálu neprobíhá.")
    else:
        best, second, worst = ranked[0], ranked[1], ranked[-1]
        close = ((by_ticker[best]["r1"] or 0) - (by_ticker[second]["r1"] or 0)) < 1.0
        target = (f"{decl[best][1]} a {decl[second][1]}" if close else decl[best][1])
        s_rotation = (f"V posledních čtyřech týdnech se relativní síla přesouvá "
                      f"z {decl[worst][0]} směrem k {target}.")

    # 3) nejsilnější sektory (tabulka je už seřazená podle 3M)
    top3 = [r["name"].lower() for r in sectors["rows"][:3]]
    s_sectors = f"Nejlepší sektorové momentum mají {top3[0]}, {top3[1]} a {top3[2]}."

    # 4) nejúspěšnější momentum ETF podle 6M
    cands = [r for r in etfs["rows"] if not r["benchmark"] and r["r6"] is not None]
    best_etf = max(cands, key=lambda r: r["r6"])
    s_etfs = (f"Z momentum ETF se nejvíce daří {best_etf['name']} "
              f"({_fmt(best_etf['r6'])} za šest měsíců).")

    # 5) chytré peníze vs retail (poslední hodnota + posun za ~4 týdny)
    def last_and_delta(chart, idx=0):
        vals = chart["series"][idx]["values"]
        return vals[-1], vals[-1] - (vals[-5] if len(vals) >= 5 else vals[0])

    # věta se skládá z indikátorů, které jsou zrovna k dispozici
    trend = lambda d, up, down, flat, lim=1: (up if d > lim else down if d < -lim else flat)
    parts = []
    if smart.get("naaim"):
        naaim, d_naaim = last_and_delta(smart["naaim"])
        parts.append(
            f"aktivní správci drží akciovou expozici {naaim:.0f} % "
            f"({trend(d_naaim, 'a za poslední měsíc ji zvyšují', 'a za poslední měsíc ji snižují', 'beze změny za poslední měsíc', 3)})"
        )
    if smart.get("retail"):
        _, d_retail = last_and_delta(smart["retail"])
        parts.append(f"apetit retailu podle objemu pákových ETF {trend(d_retail, 'roste', 'klesá', 'stagnuje')}")
    if smart.get("cot"):
        _, d_inst = last_and_delta(smart["cot"], 0)
        parts.append(f"instituce ve futures na S&P 500 {trend(d_inst, 'pozice přidávají', 'pozice ubírají', 'pozice drží')}")
    # VIX zmiňujeme jen při skutečném stresu (backwardace termínové struktury)
    if smart.get("vix"):
        term = smart["vix"]["term"]["series"][0]["values"]
        if term and term[-1] is not None and term[-1] < 1:
            parts.append("opční trh signalizuje akutní stres (termínová struktura VIXu je převrácená)")
    s_smart = (parts[0][0].upper() + ", ".join(parts)[1:] + ".") if parts else None

    # 6) Bitcoin jako barometr rizikového apetitu (signál vs 52t průměr)
    s_btc = None
    btc = next((r for r in crypto["rows"] if r["ticker"] == "BTC-USD"), None)
    if btc and btc["signal"]:
        b3 = btc["r3"]
        if btc["signal"] == "above":
            s_btc = ("Bitcoin, barometr rizikového apetitu, se drží nad svým "
                     "52týdenním průměrem"
                     + (f" ({_fmt(b3)} za tři měsíce)." if b3 is not None else "."))
        elif b3 is not None and b3 > 0:
            s_btc = (f"Bitcoin, barometr rizikového apetitu, je stále pod svým "
                     f"52týdenním průměrem, krátkodobé momentum ale sílí "
                     f"({_fmt(b3)} za tři měsíce).")
        else:
            s_btc = ("Bitcoin, barometr rizikového apetitu, je pod svým "
                     "52týdenním průměrem"
                     + (f" a za tři měsíce ztrácí {_fmt(b3)}." if b3 is not None else "."))

    # 7) likvidita – čistá likvidita Fedu + peníze vs inflace
    try:
        liq = json.loads((OUT_DIR / "liquidity.json").read_text(encoding="utf-8"))
    except Exception:
        liq = {}
    liq_parts = []
    if liq.get("net_liquidity"):
        v = liq["net_liquidity"]["series"][0]["values"]
        d12 = (v[-1] - v[-13]) if len(v) >= 13 else 0  # posun za ~3 měsíce (týdenní data)
        liq_parts.append("čistá likvidita ve finančním systému USA za poslední tři měsíce "
                         + ("roste" if d12 > 0.05 else "klesá" if d12 < -0.05 else "stagnuje"))
    if liq.get("money_us"):
        m2v = liq["money_us"]["series"][0]["values"][-1]
        cpiv = liq["money_us"]["series"][1]["values"][-1]
        if m2v is not None and cpiv is not None:
            gap = m2v - cpiv
            liq_parts.append("peněžní zásoba " + ("roste rychleji než inflace"
                             if gap > 0.5 else "zaostává za inflací"
                             if gap < -0.5 else "zhruba drží krok s inflací"))
    s_liq = None
    if liq_parts:
        joined = " a ".join(liq_parts)
        s_liq = joined[0].upper() + joined[1:] + "."

    # 8) rizika – věta jen při eskalaci proti poslednímu roku (na trvale
    # zvýšenou hladinu si trhy zvyknou, srovnání s věčným průměrem by ječelo pořád)
    s_risks = None
    try:
        risks = json.loads((OUT_DIR / "risks.json").read_text(encoding="utf-8"))
        gpr = risks.get("gpr")
        if gpr:
            last, base = gpr["series"][0]["values"][-1], gpr.get("avg_1y") or gpr.get("avg")
            if last is not None and base and last > 1.5 * base:
                s_risks = (f"Geopolitické napětí eskaluje nad úroveň posledního roku "
                           f"(index GPR {last:.0f} vs ~{base:.0f}, na které si trhy zvykly).")
    except Exception:
        pass

    (OUT_DIR / "summary.json").write_text(json.dumps({
        "updated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "sentences": {
            "global": s_global,
            "rotation": s_rotation,
            "sectors": s_sectors,
            "etfs": s_etfs,
            "smart": s_smart,
            "btc": s_btc,
            "liquidity": s_liq,
            "risks": s_risks,
        },
    }, ensure_ascii=False, indent=1), encoding="utf-8")
    print("[summary] -> summary.json")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for key, cfg in GROUPS.items():
        build_group(key, cfg)
    build_smart_money()
    build_social()
    build_liquidity()
    build_risks()
    build_polymarket()
    build_summary()
    print(f"Hotovo. Vygenerováno do {OUT_DIR}")


def main_daily() -> None:
    """Denní běh: jen Sítě (Reddit + StockTwits) – jediná data s denní
    vypovídací hodnotou. Momentum zůstává týdenní záměrně (denní přepočet
    částečných týdenních barů je šum, viz metodika)."""
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    build_social()
    print(f"Hotovo (denní běh). Vygenerováno do {OUT_DIR}")


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("--daily", action="store_true",
                    help="aktualizovat jen denní data (Sítě)")
    if ap.parse_args().daily:
        main_daily()
    else:
        main()
